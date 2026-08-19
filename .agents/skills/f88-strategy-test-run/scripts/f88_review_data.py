#!/usr/bin/env python3
"""f88_review_data.py — F88 审核造数高频操作脚本（脚本优先，浏览器兜底）。

封装 F88 预发环境三类高频 API 操作，借用已登录 Chrome 的 session（经 web-automation
的 CDPClient 在 pre-aifashion-xiaoer 页面上下文执行 fetch）：

  trial-run     策略试运行造数（方式一，首选）：xlsx/JSON → inputDatas → BT_批次
  batch-status  查询/轮询批次状态（getRunDetail）
  strategy-info 查询策略 inputParams + 节点编排
  create-task   手动创建审核任务（方式二，仅 formal 语义验证用，绕过 workflow 管线）

前置条件：
  1. Chrome 以调试模式运行且已登录 pre-aifashion-xiaoer.alibaba-inc.com：
     /Applications/Google\\ Chrome.app/Contents/MacOS/Google\\ Chrome \\
       --remote-debugging-port=9223 --user-data-dir="$HOME/.config/chrome-debug"
  2. web-automation skill 已安装 puppeteer-core：
     cd ~/.qoderwork/skills/web-automation && npm install
  3. xlsx 输入需要 openpyxl（pip3 install openpyxl）；也可用 --json 直接传 payload

示例：
  python3 f88_review_data.py trial-run --strategy 10834 --mode test \\
      --xlsx /Users/caoxuemei/qoder/f88素材生产/审核专用模板.xlsx
  python3 f88_review_data.py batch-status --batch BT_7544 --watch
  python3 f88_review_data.py strategy-info --id 10834
  python3 f88_review_data.py create-task --node 168 --data-file-url https://oss/...

退出码：0 成功；1 环境/前置错误；2 API 返回失败。
"""

import argparse
import asyncio
import base64
import json
import os
import sys
import time

WEB_AUTO_DIR = os.path.expanduser("~/.qoderwork/skills/web-automation")
TARGET_HOST = "pre-aifashion-xiaoer.alibaba-inc.com"
TARGET_URL = f"https://{TARGET_HOST}/"
IDENTITY_HEADER = "f88"  # X-AFD-Emp-Identity，所有 /api/afd/ 与 workflow 请求必带

KNOWN_COLUMNS = [
    "seller_id", "seed_image_url", "main_img_url", "fabric_tryon_url",
    "fabric_url", "item_id", "tao_cate",
]


def die(msg, code=1):
    print(f"[ERROR] {msg}", file=sys.stderr)
    sys.exit(code)


def load_cdp_client():
    if not os.path.isdir(WEB_AUTO_DIR):
        die(f"web-automation skill 不存在：{WEB_AUTO_DIR}")
    if not os.path.isdir(os.path.join(WEB_AUTO_DIR, "node_modules", "puppeteer-core")):
        die("puppeteer-core 未安装，请先执行：cd %s && npm install" % WEB_AUTO_DIR)
    sys.path.insert(0, WEB_AUTO_DIR)
    try:
        from core.cdp_client import CDPClient
    except Exception as e:  # noqa: BLE001
        die(f"导入 CDPClient 失败：{e}")
    return CDPClient


def build_fetch_js(path, method="GET", body=None, extra_headers=None):
    """构造页面内 fetch 的 async IIFE JS 表达式。"""
    headers = {"X-AFD-Emp-Identity": IDENTITY_HEADER}
    if body is not None:
        headers["Content-Type"] = "application/json"
    if extra_headers:
        headers.update(extra_headers)
    opts = {
        "method": method,
        "credentials": "include",
        "headers": headers,
    }
    if body is not None:
        opts["body"] = json.dumps(body, ensure_ascii=False)
    return (
        "(async () => { try {"
        f" const resp = await fetch({json.dumps(path)}, {json.dumps(opts, ensure_ascii=False)});"
        " const data = await resp.json().catch(() => null);"
        " return { status: resp.status, data };"
        " } catch (e) { return { status: 0, error: String(e && e.message || e) }; } })()"
    )


async def api(cdp, path, method="GET", body=None):
    try:
        result = await asyncio.wait_for(cdp.evaluate(build_fetch_js(path, method, body)), timeout=60)
    except asyncio.TimeoutError:
        die(f"API 超时（60s）：{path}")
    if not isinstance(result, dict):
        die(f"API 返回异常：{result!r}")
    if result.get("status") == 0:
        die(f"fetch 执行失败：{result.get('error')}")
    if result.get("status") != 200:
        die(f"HTTP {result.get('status')}：{path} → {json.dumps(result.get('data'), ensure_ascii=False)[:500]}")
    return result.get("data")


def read_xlsx_rows(path):
    try:
        import openpyxl
    except ImportError:
        die("缺少 openpyxl，请 pip3 install openpyxl，或改用 --json 传入 inputDatas")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {k: str(v) for k, v in zip(headers, row) if k and v is not None and str(v).strip() != ""}
        if d:
            rows.append(d)
    if not rows:
        die(f"xlsx 无数据行：{path}")
    return rows


def filter_by_params(rows, codes):
    """按策略 inputParams 的 code 过滤列；未知列丢弃，缺列告警。"""
    out, warned = [], False
    for r in rows:
        filtered = {k: v for k, v in r.items() if k in codes}
        missing = [c for c in codes if c not in filtered]
        if missing and not warned:
            print(f"[WARN] 输入缺少策略要求的列：{missing}", file=sys.stderr)
            warned = True
        if filtered:
            out.append(filtered)
    return out


async def connect(cdp_cls):
    cdp = cdp_cls()
    try:
        await cdp.connect(url_pattern=TARGET_HOST, url=TARGET_URL)
    except Exception as e:  # noqa: BLE001
        die(f"连接 Chrome 失败：{e}\n请确认 Chrome 已以 --remote-debugging-port=9223 启动（详见脚本头部说明）")
    try:
        login = await cdp.check_login()
        if isinstance(login, dict) and login.get("isLoginPage"):
            die("浏览器未登录（被重定向到登录页），请先在 Chrome 中手动登录 pre-aifashion-xiaoer")
    except Exception:  # noqa: BLE001 check_login 非关键路径
        pass
    return cdp


# ---------------- subcommands ----------------

async def cmd_trial_run(args):
    cdp = await connect(load_cdp_client())
    try:
        # 1. 查策略 inputParams
        info = await api(cdp, f"/api/workflow2/strategy/get?id={args.strategy}")
        data = (info or {}).get("data") or {}
        wf = data.get("workflowDef") or {}
        codes = [p.get("code") for p in (wf.get("inputParams") or []) if p.get("code")]
        if not codes:
            die(f"策略 {args.strategy} 未返回 inputParams，请确认策略 ID 与发布状态")
        print(f"[INFO] 策略：{data.get('name')}，inputParams={codes}", file=sys.stderr)

        # 2. 组装 inputDatas
        if args.json:
            with open(args.json, encoding="utf-8") as f:
                rows = json.load(f)
        else:
            rows = read_xlsx_rows(args.xlsx)
        input_datas = filter_by_params(rows, codes)
        if args.rows and args.rows > 0:
            input_datas = input_datas[: args.rows]
        print(f"[INFO] inputDatas 共 {len(input_datas)} 行，runMode={args.mode}", file=sys.stderr)

        # 3. 触发试运行
        resp = await api(cdp, "/api/workflow2/strategy/run", method="POST",
                         body={"strategyId": args.strategy, "inputDatas": input_datas, "runMode": args.mode})
        if not (resp or {}).get("success"):
            die(f"试运行失败：{json.dumps(resp, ensure_ascii=False)[:500]}", code=2)
        batch_id = resp.get("data")
        print(json.dumps({"success": True, "batchId": batch_id, "strategyId": args.strategy,
                          "runMode": args.mode, "rowCount": len(input_datas)}, ensure_ascii=False))

        # 4. 轮询批次状态
        if not args.no_poll and batch_id:
            await poll_batch(cdp, batch_id, timeout=args.poll_timeout)
    finally:
        await cdp.disconnect()


async def poll_batch(cdp, batch_id, timeout=180, interval=10):
    deadline = time.time() + timeout
    last = None
    while time.time() < deadline:
        resp = await api(cdp, f"/api/workflow/batch/getRunDetail?batchId={batch_id}")
        wb = ((resp or {}).get("data") or {}).get("workflowBatch") or {}
        status = wb.get("status")
        if status != last:
            print(f"[POLL] {batch_id} status={status}", file=sys.stderr)
            last = status
        if status in ("SUCCESS", "FAIL", "FINISH"):
            print(json.dumps({"batchId": batch_id, "status": status,
                              "relationId": wb.get("relationId")}, ensure_ascii=False))
            return
        await asyncio.sleep(interval)
    print(f"[POLL] {timeout}s 内未到终态（最后状态 {last}）。BATCH 模式需等 ScheduleX 攒批周期，"
          "可稍后用 batch-status --watch 复查；勿误判为卡死。", file=sys.stderr)


async def cmd_batch_status(args):
    cdp = await connect(load_cdp_client())
    try:
        if args.watch:
            await poll_batch(cdp, args.batch, timeout=args.poll_timeout)
        else:
            resp = await api(cdp, f"/api/workflow/batch/getRunDetail?batchId={args.batch}")
            print(json.dumps(resp, ensure_ascii=False, indent=2))
    finally:
        await cdp.disconnect()


async def cmd_strategy_info(args):
    cdp = await connect(load_cdp_client())
    try:
        info = await api(cdp, f"/api/workflow2/strategy/get?id={args.id}")
        data = (info or {}).get("data") or {}
        wf = data.get("workflowDef") or {}
        out = {
            "strategyId": args.id,
            "name": data.get("name"),
            "inputParams": [p.get("code") for p in (wf.get("inputParams") or [])],
            "innerNodes": [
                {"name": n.get("name"), "approveNodeId": n.get("approveNodeId"),
                 "approveType": n.get("approveType"), "execMode": n.get("execMode")}
                for n in (wf.get("innerNodes") or [])
            ],
        }
        print(json.dumps(out, ensure_ascii=False, indent=2))
    finally:
        await cdp.disconnect()


async def cmd_create_task(args):
    """方式二：手动创建审核任务（不推荐，仅 formal 语义验证）。"""
    cdp = await connect(load_cdp_client())
    try:
        data_file_url = args.data_file_url
        if not data_file_url:
            if not args.upload_xlsx:
                die("需要 --data-file-url 或 --upload-xlsx 之一")
            data_file_url = await upload_file(cdp, args.upload_xlsx)
            print(f"[INFO] 上传成功：{data_file_url}", file=sys.stderr)

        body = {
            "taskName": args.task_name,
            "nodeId": args.node,
            "dataFileUrl": data_file_url,
            "standardIds": [args.standard_id],
            "priority": 0,
            "expectedDeliveryTime": args.delivery_time,
            "difficulty": 2,
            "efficiency": 500,
            "allocation": {
                "roles": ["reviewer"],
                "requiredTagIds": [],
                "participants": [{"userId": args.reviewer_id, "userName": args.reviewer_name,
                                  "count": args.count}],
                "allocationMethod": 2,
            },
            "inspectionConfig": {"enabled": False, "participantUserIds": [], "distributionType": 1,
                                 "sampleSourceUserIds": [], "ratio": 0, "maxCountPerUser": 0,
                                 "perPersonCount": 0},
            "buryConfig": {"enabled": False, "ratio": 0, "maxCountPerUser": 0, "perPersonCount": 0},
            "distributionLogic": 1,
        }
        resp = await api(cdp, "/api/afd/review/task/main/create", method="POST", body=body)
        if not (resp or {}).get("success"):
            die(f"创建失败：{json.dumps(resp, ensure_ascii=False)[:500]}", code=2)
        task_id = resp.get("data")
        print(json.dumps({"success": True, "taskId": task_id, "nodeId": args.node,
                          "dataFileUrl": data_file_url}, ensure_ascii=False))
        # 回查验证
        detail = await api(cdp, f"/api/afd/review/task/main/parentReviewTaskDetail?taskId={task_id}")
        ok = bool(((detail or {}).get("data")))
        print(json.dumps({"verified": ok,
                          "note": "确认 dataFileUrl 非 null 且 UI 图片正常；DB 验证走 dms-alibaba"},
                         ensure_ascii=False))
    finally:
        await cdp.disconnect()


async def upload_file(cdp, path):
    """经页面内 /api/file/upload 上传本地文件，返回 OSS URL。"""
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    name = os.path.basename(path)
    js = (
        "(async () => { try {"
        f" const bin = atob({json.dumps(b64)});"
        " const bytes = new Uint8Array(bin.length);"
        " for (let i = 0; i < bin.length; i++) bytes[i] = bin.charCodeAt(i);"
        f" const blob = new Blob([bytes]);"
        " const fd = new FormData();"
        f" fd.append('file', blob, {json.dumps(name)});"
        " const resp = await fetch('/api/file/upload', { method: 'POST', credentials: 'include',"
        "   headers: { 'X-AFD-Emp-Identity': 'f88' }, body: fd });"
        " const data = await resp.json().catch(() => null);"
        " return { status: resp.status, data };"
        " } catch (e) { return { status: 0, error: String(e && e.message || e) }; } })()"
    )
    result = await asyncio.wait_for(cdp.evaluate(js), timeout=120)
    if not isinstance(result, dict) or result.get("status") != 200:
        die(f"文件上传失败：{json.dumps(result, ensure_ascii=False)[:300]}")
    url = ((result.get("data") or {}).get("data"))
    if not url:
        die(f"上传响应无 OSS URL：{json.dumps(result.get('data'), ensure_ascii=False)[:300]}")
    return url


def main():
    ap = argparse.ArgumentParser(description="F88 审核造数脚本（脚本优先，浏览器兜底）")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("trial-run", help="策略试运行造数（方式一，首选）")
    p.add_argument("--strategy", type=int, required=True, help="策略 ID，如 10834(流式)/10833(块式)")
    p.add_argument("--mode", choices=["test", "formal"], default="test",
                   help="test=链路验证；formal=过滤规则(自过滤等)验证")
    p.add_argument("--xlsx", help="输入 xlsx（默认审核专用模板.xlsx）")
    p.add_argument("--json", help="直接提供 inputDatas JSON 文件（优先于 --xlsx）")
    p.add_argument("--rows", type=int, default=0, help="只取前 N 行（0=全部）")
    p.add_argument("--no-poll", action="store_true", help="触发后不轮询批次状态")
    p.add_argument("--poll-timeout", type=int, default=180, help="轮询超时秒数（默认 180）")

    p = sub.add_parser("batch-status", help="查询/轮询批次状态")
    p.add_argument("--batch", required=True, help="批次号 BT_xxx")
    p.add_argument("--watch", action="store_true", help="轮询直到终态或超时")
    p.add_argument("--poll-timeout", type=int, default=180)

    p = sub.add_parser("strategy-info", help="查询策略 inputParams 与节点编排")
    p.add_argument("--id", type=int, required=True)

    p = sub.add_parser("create-task", help="手动创建审核任务（方式二，仅 formal 语义验证）")
    p.add_argument("--node", type=int, required=True, help="168=首图 139=套图 144=视频 138=模板")
    p.add_argument("--data-file-url", help="已有数据文件 OSS URL")
    p.add_argument("--upload-xlsx", help="本地 xlsx，脚本自动上传取 OSS URL")
    p.add_argument("--task-name", default="自动化测试任务")
    p.add_argument("--standard-id", type=int, default=140)
    p.add_argument("--delivery-time", default="2026-12-31 12:00:00")
    p.add_argument("--reviewer-id", default="526043", help="审核人工号（红线：必须为目民 526043）")
    p.add_argument("--reviewer-name", default="目民")
    p.add_argument("--count", type=int, default=5)

    args = ap.parse_args()
    if args.cmd == "trial-run" and not args.xlsx and not args.json:
        args.xlsx = "/Users/caoxuemei/qoder/f88素材生产/审核专用模板.xlsx"
        if not os.path.exists(args.xlsx):
            die(f"默认模板不存在：{args.xlsx}，请用 --xlsx 或 --json 指定输入")

    handlers = {
        "trial-run": cmd_trial_run,
        "batch-status": cmd_batch_status,
        "strategy-info": cmd_strategy_info,
        "create-task": cmd_create_task,
    }
    asyncio.run(handlers[args.cmd](args))


if __name__ == "__main__":
    main()
